import openpyxl
import csv
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

'''Чтение csv-файлов и конвертация в excel-формат'''
# wb = openpyxl.Workbook()
# ws = wb.active
#
# with open('TuTu_Flights.csv', encoding='utf8') as f:
#     reader = csv.reader(f, delimiter=',')
#     for row in reader:
#         ws.append(row)
#
# wb.save('TuTu_Flights.xlsx')

'''
1. Создаём список с названиями эксель-файлов.
2. Создаём список заголовков колонок, подлежащих форматированию.
3. Проходимся циклом по этим файлам. 
4. Определяем длину самого длинного значения ячейки в колонках и расширяем эту колонку на 7 символов.
5. Создаём объект класса Border для создания границ ячеек.
6. Выделяем первую строку (заголовки) полужирным шрифтом.
7. Выравниваем значения ячеек по центру.
8. Создаём словарь, где ключ - заголовок, а значение - порядковая буква соответствующей колонки.
9. У ячеек в колонках "Минимальная цена билета", "Время вылета/прилёта" инициализируем переменные cell.fill 
и присваиваем им значения кодировок соответствующих цветов; цены дополнительно выделяем жирным шрифтом.
10. Определяем диапазон таблицы и создаём автофильтры для каждой колонки.'''

docs = ['Aviasales_Flights.xlsx', 'TuTu_Flights.xlsx']

headers_of_interest = ["Минимальная цена билета", "Время вылета", "Время прилёта"]

for doc in docs:
    avia_book = openpyxl.load_workbook(doc)
    sheet = avia_book['Sheet']

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 7

    thin_border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.border = thin_border

    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for column in sheet.columns:
        for cell in column:
            cell.alignment = alignment
            cell.border = thin_border

    header_to_column = {cell.value: cell.column_letter for cell in sheet[1]}

    light_orange_fill = PatternFill(start_color='E3CFC3', end_color='E3CFC3', fill_type='solid')
    light_green_fill = PatternFill(start_color='CFE4B7', end_color='CFE4B7', fill_type='solid')

    for header in headers_of_interest:
        if header in header_to_column:
            column_letter = header_to_column[header]
            for cell in sheet[column_letter]:
                if header == "Минимальная цена билета":
                    cell.fill = light_orange_fill
                    cell.font = Font(bold=True)
                elif header in ["Время вылета", "Время прилёта"]:
                    cell.fill = light_green_fill
                cell.border = thin_border

    last_column_letter = sheet.cell(row=1, column=sheet.max_column).column_letter
    auto_filter_range = f"A1:{last_column_letter}{sheet.max_row}"

    sheet.auto_filter.ref = auto_filter_range

    avia_book.save(doc)
